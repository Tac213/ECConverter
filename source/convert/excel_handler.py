# -*- coding: utf-8 -*-
# author: Tac
# contact: cookiezhx@163.com

import math
import typing

import openpyxl
from openpyxl.worksheet import worksheet

import genv
from const import convert_const


def get_workbook(excel_path: str) -> openpyxl.Workbook:
    """
    读取Excel的工作簿, 而且只读取数据
    Args:
        excel_path: [str]excel完整路径
    Returns:
        openpyxl.workbook.Workbook
    """
    return openpyxl.load_workbook(excel_path, read_only=False, data_only=True)


def get_sheets(workbook: openpyxl.Workbook) -> typing.List[worksheet.Worksheet]:
    """
    读取工作簿的所有sheet
    Args:
        workbook: openpyxl.workbook.Workbook
    Returns:
        list, 每一项是openpyxl.worksheet.worksheet.Worksheet
    """
    return workbook.worksheets


def get_sheet_by_name(workbook: openpyxl.Workbook, sheet_name: str) -> worksheet.Worksheet:
    """
    根据sheet名字获取sheet
    Args:
        workbook: openpyxl.workbook.Workbook
        sheet_name: [str]sheet名字
    Returns:
        openpyxl.worksheet.worksheet.Worksheet
    """
    return workbook[sheet_name]


def get_data_name(sheet: worksheet.Worksheet) -> str:
    """
    获取某一页sheet对应的data名字
    Args:
        sheet: openpyxl.worksheet.worksheet.Worksheet
    Returns:
        str
    """
    data_name = get_cell_value_str(sheet, *genv.settings.data_name_coordinate)
    if genv.settings.data_name_splitter in data_name:
        if not data_name.split(genv.settings.data_name_splitter)[0].endswith(genv.settings.data_name_suffix):
            return None
    elif not data_name.endswith(genv.settings.data_name_suffix):
        return None
    return data_name


def get_enum_class_name(sheet: worksheet.Worksheet) -> str:
    """
    获取某一页sheet对应的enum类名
    Args:
        sheet: openpyxl.worksheet.worksheet.Worksheet
    Returns:
        str
    """
    data_name = get_cell_value_str(sheet, *genv.settings.data_name_coordinate)
    return data_name


def get_row_generator(sheet: worksheet.Worksheet, start_idx: int) -> typing.Generator[tuple, None, None]:
    """
    获取行的生成器, 用于遍历sheet
    Args:
        sheet: openpyxl.worksheet.worksheet.Worksheet
        start_idx: [int]起始索引值, 索引值从0开始
    Returns:
        types.GeneratorType
    """
    return sheet.iter_rows(min_row=start_idx + 1, values_only=True)


def get_col_generator(sheet: worksheet.Worksheet, start_idx: int) -> typing.Generator[tuple, None, None]:
    """
    获取列的生成器, 用于遍历sheet
    Args:
        sheet: openpyxl.worksheet.worksheet.Worksheet
        start_idx: [int]起始索引值, 索引值从0开始
    Returns:
        types.GeneratorType
    """
    return sheet.iter_cols(min_col=start_idx + 1, values_only=True)


def get_sheet_name(sheet: worksheet.Worksheet) -> str:
    """
    获取sheet名字
    Args:
        sheet: openpyxl.worksheet.worksheet.Worksheet
    Returns:
        str
    """
    return sheet.title


def get_sheet_row_count(sheet: worksheet.Worksheet) -> int:
    """
    获取sheet的行数
    Args:
        sheet: openpyxl.worksheet.worksheet.Worksheet
    Returns:
        int
    """
    return sum(1 for _ in sheet.rows)


def get_sheet_column_count(sheet: worksheet.Worksheet) -> int:
    """
    获取sheet的行数
    Args:
        sheet: openpyxl.worksheet.worksheet.Worksheet
    Returns:
        int
    """
    return sum(1 for _ in sheet.columns)


def get_cell_value(sheet: worksheet.Worksheet, row_idx: int, col_idx: int) -> typing.Any:
    """
    获取单元格数据
    Args:
        sheet: openpyxl.worksheet.worksheet.Worksheet
        row_idx: [int]行索引值, 索引值从0开始
        col_idx: [int]列索引值, 索引值从0开始
    Returns:
        sheet原本的数据
    """
    return sheet.cell(row_idx + 1, col_idx + 1).value


def get_value_str(value: typing.Any) -> str:
    """
    将excel原始数据转为字符串
    Args:
        value: excel数据
    Returns:
        str, 经过字符串处理的数据
    """
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.fabs(value - int(value)) < convert_const.INT_THRESHOLD:
            value = str(int(value))
        else:
            value = str(value)
        return value
    if value is None:
        return ''
    return str(value).strip()


def get_cell_value_str(sheet: worksheet.Worksheet, row_idx: int, col_idx: int) -> str:
    """
    获取经过字符串处理的单元格数据
    Args:
        sheet: openpyxl.worksheet.worksheet.Worksheet
        row_idx: [int]行索引值, 索引值从0开始
        col_idx: [int]列索引值, 索引值从0开始
    Returns:
        str, 经过字符串处理的单元格数据
    """
    value = get_cell_value(sheet, row_idx, col_idx)
    return get_value_str(value)


def get_cell_value_str_with_tuple(row_or_col_data_tuple: tuple, idx: int) -> str:
    """
    通过生成器得到的行数据或列数据元组，获取经过字符串处理的单元格数据
    Args:
        row_or_col_data_tuple: [tuple]通过生成器得到的行数据或列数据元组
        idx: [int]索引值, 从0开始
    Returns:
        str, 经过字符串处理的单元格数据
    """
    value = row_or_col_data_tuple[idx]
    return get_value_str(value)


def get_column_name(col_idx: int) -> str:
    """
    根据列索引值获取列名
    Args:
        col_idx: [int]列索引值, 索引值从0开始
    Returns:
        str
    """
    return openpyxl.utils.get_column_letter(col_idx + 1)


def get_cell_name(row_idx: int, col_idx: int) -> str:
    """
    获取单元格名字
    Args:
        row_idx: [int]行索引值, 索引值从0开始
        col_idx: [int]列索引值, 索引值从0开始
    Returns:
        str
    """
    return f'{get_column_name(col_idx)}{row_idx + 1}'


def set_column_validator_type_list(sheet: worksheet.Worksheet, col_idx: int, selections: list, row_offset: int) -> bool:
    """
    设置worksheet某一列的validator为下拉列表
    Args:
        worksheet: openpyxl.worksheet.worksheet.Worksheet
        col_idx: [int]从0开始的列索引值
        selections: [list]选项列表
        row_offset: [int]行偏移值, 也可以理解成开始行的索引值, 索引值从0开始
    Returns:
        bool, 是否成功设置
    """
    flabel = get_cell_value_str(sheet, genv.settings.field_label_index, col_idx)
    if flabel == convert_const.ValidFieldLabel.REPEATED:
        return False
    formula_str = f"\"{','.join(map(str, selections))}\""
    data_validator = openpyxl.worksheet.datavalidation.DataValidation(type='list', formula1=formula_str, allow_blank=True)
    col_letter = get_column_name(col_idx)
    start_cell_name = f'{col_letter}{row_offset + 1}'
    end_cell_name = f'{col_letter}{get_sheet_row_count(sheet)}'
    data_validator.add(f'{start_cell_name}:{end_cell_name}')
    sheet.add_data_validation(data_validator)
    return True


def set_worksheet_cell_value(sheet: worksheet.Worksheet, row_idx: int, col_idx: int, value: typing.Any) -> bool:
    """
    写入单元格的值
    Args:
        worksheet: openpyxl.worksheet.worksheet.Worksheet
        row_idx: [int]从0开始的行索引值
        col_idx: [int]从0开始的列索引值
        value: 写入的值
    Returns:
        bool是否成功写入
    """
    sheet.cell(row=row_idx + 1, column=col_idx + 1).value = value
    return True


def set_worksheet_cell_comment(sheet: worksheet.Worksheet,
                               row_idx: int,
                               col_idx: int,
                               comment_str: str,
                               author: str,
                               width: typing.Optional[float] = None,
                               height: typing.Optional[float] = None) -> bool:
    """
    写入单元格的注释
    Args:
        worksheet: openpyxl.worksheet.worksheet.Worksheet
        row_idx: [int]从0开始的行索引值
        col_idx: [int]从0开始的列索引值
        comment_str: [str]注释内容
        author: [str]注释作者
        width: [float]注释框宽度, 传None会根据comment的行宽算出来
        height: [float]注释框高度, 传None会根据comment的行数算出来
    Returns:
        bool是否成功写入
    """
    comment = openpyxl.comments.Comment(f'{author}:\n{comment_str}', author)
    if not width or not height:
        comment_split = comment_str.split('\n')
        max_row_length = max(map(len, comment_split))
        row_cnt = len(comment_split) + 1  # 加上author那一行
        if not width:
            width = max_row_length * convert_const.COMMENT_CHAR_WIDTH
        if not height:
            height = row_cnt * convert_const.COMMENT_CHAR_HEIGHT
    comment.width = width
    comment.height = height
    sheet.cell(row=row_idx + 1, column=col_idx + 1).comment = comment
    return True
